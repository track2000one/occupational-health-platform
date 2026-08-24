import hashlib
import json
import re
from datetime import datetime, date
from typing import Any, Dict, Iterable, List, Optional, Tuple

from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .models import (
    CampaignCoverage,
    ClinicVisit,
    Employee,
    EmployeeImportReview,
    EmployeeClinicVisit,
    EmployeeHealthProfile,
    HealthCenter,
    LabScreening,
    MedicalCommitteeCase,
    NeedleStickExposure,
    OccupationalClinicVisit,
    ReferenceLookup,
    ScreeningProgramRecord,
    TraineeRotation,
    VaccinationRecord,
)

SENSITIVE_FIELDS = {'national_id', 'mobile', 'phone', 'source_national_id'}
DEFAULT_CENTER_NAME = 'Imported Occupational Health Center'

HEADER_ALIASES = {
    'name': ['name', 'employee name', 'full name', 'اسم', 'الأسم', 'الاسم', 'اسم الموظف', 'employee', 'الموظف'],
    'national_id': ['national id', 'national_id', 'nid', 'id', 'رقم الهوية', 'رقم الهوية /الاقامة', 'الهوية', 'الهويه', 'رقم السجل', 'السجل المدني', 'السج المدني', 'رقم الهوية الوطنية', 'national id'],
    'mobile': ['mobile', 'mobile #', 'phone', 'phone number', 'جوال', 'الجوال', 'رقم الجوال', 'الهاتف', 'رقم الهاتف'],
    'gender': ['gender', 'sex', 'الجنس'],
    'health_center': ['health center', 'center', 'facility', 'المركز', 'المركز الصحي', 'الجهة', 'جهة العمل', 'المستشفى'],
    'job_title': ['job title', 'position', 'job', 'مسمى الوظيفة', 'المسمى الوظيفي', 'الوظيفة', 'المهنة'],
    'age': ['age', 'العمر'],
    'diagnosis': ['diagnosis', 'diagnoses', 'التشخيص', 'تشخيص'],
    'clinic_type': ['clinic', 'clinic type', 'العيادة', 'نوع العيادة', 'نوع الزيارة'],
    'action_taken': ['action', 'action taken', 'recommendation', 'recommendations', 'الإجراء', 'الاجراء', 'التوصية', 'التوصيات'],
    'visit_date': ['date', 'visit date', 'تاريخ', 'التاريخ', 'تاريخ الزيارة'],
}

EMPLOYEE_ROSTER_ALIASES = {
    'name': ['name', 'employee name', 'full name', 'staff name', 'اسم الموظف', 'الاسم', 'الاسم الكامل', 'الاسم الرباعي'],
    'email': ['email', 'email address', 'e-mail', 'البريد', 'البريد الإلكتروني', 'البريد الالكتروني'],
    'national_id': ['national id', 'national_id', 'id number', 'civil id', 'رقم الهوية الوطنية', 'رقم الهوية', 'الهوية', 'السجل المدني', 'رقم السجل المدني'],
    'employee_number': ['employee number', 'employee no', 'employee id', 'staff id', 'staff number', 'الرقم الوظيفي', 'رقم الموظف', 'رقم المنسوب'],
    'national_address': ['national address', 'address', 'العنوان الوطني', 'العنوان'],
    'mobile': ['mobile', 'mobile number', 'mobile no', 'phone', 'phone number', 'رقم الجوال', 'الجوال', 'رقم الهاتف', 'الهاتف'],
    'date_of_birth': ['date of birth', 'birth date', 'dob', 'تاريخ الميلاد', 'الميلاد'],
    'birth_place': ['birth place', 'place of birth', 'مكان الميلاد', 'محل الميلاد'],
    'gender': ['gender', 'sex', 'الجنس', 'النوع'],
    'marital_status': ['marital status', 'social status', 'الحالة الاجتماعية', 'الحالة'],
    'health_center': ['health center', 'facility', 'work facility', 'المركز الصحي', 'المركز', 'المنشأة الصحية', 'المرفق الصحي', 'جهة العمل'],
    'job_title': ['job title', 'position', 'occupation', 'المسمى الوظيفي', 'الوظيفة', 'المهنة'],
    'appointment_date': ['appointment date', 'hire date', 'start date', 'joining date', 'date hired', 'تاريخ التعيين', 'تاريخ المباشرة', 'تاريخ الالتحاق'],
    'periodic_exam_status': ['periodic exam status', 'حالة الفحص الدوري'],
    'vaccination_status': ['vaccination status', 'حالة التطعيم'],
    'risk_level': ['risk level', 'مستوى الخطورة'],
}

EMPLOYEE_REQUIRED_FIELDS = (
    'name', 'email', 'national_id', 'employee_number', 'national_address', 'mobile',
    'date_of_birth', 'birth_place', 'gender', 'marital_status', 'health_center',
    'job_title', 'appointment_date',
)

EMPLOYEE_CHOICE_MAPS = {
    'gender': {
        'male': 'male', 'ذكر': 'male', 'm': 'male',
        'female': 'female', 'أنثى': 'female', 'انثى': 'female', 'f': 'female',
    },
    'marital_status': {
        'single': 'single', 'أعزب': 'single', 'اعزب': 'single', 'عزباء': 'single',
        'married': 'married', 'متزوج': 'married', 'متزوجة': 'married',
        'divorced': 'divorced', 'مطلق': 'divorced', 'مطلقة': 'divorced',
        'widowed': 'widowed', 'أرمل': 'widowed', 'ارمل': 'widowed', 'أرملة': 'widowed', 'ارملة': 'widowed',
    },
    'periodic_exam_status': {
        'completed': 'completed', 'مكتمل': 'completed',
        'incomplete': 'incomplete', 'غير مكتمل': 'incomplete',
        'overdue': 'overdue', 'متأخر': 'overdue',
    },
    'vaccination_status': {
        'completed': 'completed', 'مكتمل': 'completed',
        'due': 'due', 'مستحق': 'due',
        'refused': 'refused', 'مرفوض': 'refused',
    },
    'risk_level': {
        'low': 'low', 'منخفض': 'low',
        'medium': 'medium', 'متوسط': 'medium',
        'high': 'high', 'مرتفع': 'high',
    },
}


def normalize_header(value: Any) -> str:
    text = str(value or '').strip().lower().replace('\n', ' ').replace('\r', ' ')
    return ' '.join(text.split())


def clean_text(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def json_safe(value: Any):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def row_raw_payload(headers: List[str], row: Tuple[Any, ...]) -> Dict[str, Any]:
    payload: Dict[str, Any] = {}
    for idx, value in enumerate(row):
        if idx >= len(headers):
            continue
        header = clean_text(headers[idx]) or f'column_{idx + 1}'
        if header in payload:
            header = f'{header}_{idx + 1}'
        payload[header] = json_safe(value)
    return payload


def mask_value(value: Any) -> str:
    value = clean_text(value)
    if not value:
        return ''
    if len(value) <= 4:
        return '****'
    return f"{'*' * max(len(value) - 4, 4)}{value[-4:]}"


def parse_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ''):
            return default
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def parse_decimal(value: Any):
    try:
        if value in (None, ''):
            return None
        return round(float(str(value).strip()), 2)
    except (TypeError, ValueError):
        return None


def parse_bool(value: Any):
    text = normalize_header(value)
    if not text:
        return None
    if text in ('yes', 'y', 'true', '1', 'done', 'positive', 'نعم', 'تم', 'محول', 'معروف'):
        return True
    if text in ('no', 'n', 'false', '0', 'not done', 'negative', 'لا', 'غير معروف'):
        return False
    return None


def parse_gender(value: Any) -> str:
    text = normalize_header(value)
    if text in ('female', 'f', 'أنثى', 'انثى'):
        return 'female'
    return 'male'


def parse_date(value: Any):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            if 20000 <= float(value) <= 80000:
                parsed = from_excel(value)
                return parsed.date() if isinstance(parsed, datetime) else parsed
        except Exception:
            return None
    text = clean_text(value)
    if not text:
        return None
    text = text.split()[0]
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return parsed if parsed.year >= 1900 else None
        except ValueError:
            continue
    return None


def find_header_row(ws) -> Tuple[int, List[str]]:
    best_row = 1
    best_headers: List[str] = []
    best_score = 0
    detected_max_row = ws.max_row or 20
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(detected_max_row, 20), values_only=True), start=1):
        headers = [clean_text(cell) for cell in row]
        normalized = [normalize_header(cell) for cell in headers]
        non_empty = [cell for cell in normalized if cell]
        score = len(non_empty)
        for aliases in HEADER_ALIASES.values():
            if any(alias in normalized for alias in aliases):
                score += 3
        if score > best_score and len(non_empty) >= 3:
            best_score = score
            best_row = idx
            best_headers = headers
    if not best_headers:
        raise ValueError('Could not detect a valid header row in the selected sheet.')
    return best_row, best_headers


def col_index(headers: List[str], aliases: Iterable[str], after: Optional[int] = None) -> Optional[int]:
    normalized = [normalize_header(header) for header in headers]
    normalized_aliases = [normalize_header(alias) for alias in aliases]
    start = 0 if after is None else after + 1
    for idx in range(start, len(normalized)):
        if normalized[idx] in normalized_aliases:
            return idx
    return None


def val(row: Tuple[Any, ...], headers: List[str], aliases: Iterable[str], default: Any = '', after: Optional[int] = None):
    idx = col_index(headers, aliases, after=after)
    if idx is None or idx >= len(row):
        return default
    return row[idx]


def text_val(row: Tuple[Any, ...], headers: List[str], aliases: Iterable[str], default: str = '', after: Optional[int] = None) -> str:
    return clean_text(val(row, headers, aliases, default=default, after=after))


def get_or_create_center(name: Any) -> HealthCenter:
    center_name = clean_text(name) or DEFAULT_CENTER_NAME
    center, _ = HealthCenter.objects.get_or_create(name=center_name, defaults={'city': '', 'is_active': True})
    return center


def upsert_employee(payload: Dict[str, Any]) -> Employee:
    center = get_or_create_center(payload.get('health_center'))
    employee, _ = Employee.objects.update_or_create(
        national_id=payload['national_id'],
        defaults={
            'name': payload.get('name') or payload['national_id'],
            'mobile': payload.get('mobile', ''),
            'gender': payload.get('gender') or 'male',
            'health_center': center,
            'job_title': payload.get('job_title', ''),
            'age': parse_int(payload.get('age'), 0),
            'periodic_exam_status': payload.get('periodic_exam_status', 'incomplete'),
            'vaccination_status': payload.get('vaccination_status', 'due'),
            'risk_level': payload.get('risk_level', 'low'),
        },
    )
    return employee


def masked_preview(payload: Dict[str, Any]) -> Dict[str, Any]:
    safe = {key: json_safe(value) for key, value in payload.items() if not key.startswith('_')}
    for field in SENSITIVE_FIELDS:
        if field in safe:
            safe[field] = mask_value(safe.get(field, ''))
    return safe


def update_or_create_by_filter(model, filter_kwargs: Dict[str, Any], defaults: Dict[str, Any]):
    instance = model.objects.filter(**filter_kwargs).first()
    if instance:
        for key, value in defaults.items():
            setattr(instance, key, value)
        instance.save()
        return instance, False
    data = {**filter_kwargs, **defaults}
    return model.objects.create(**data), True


def make_trace(headers: List[str], row: Tuple[Any, ...], row_number: int, sheet_name: str, file_name: str) -> Dict[str, Any]:
    return {'imported_sheet': sheet_name, 'imported_row': row_number, 'source_file': file_name, 'raw_payload': row_raw_payload(headers, row)}


def iter_data_rows(ws, header_row: int):
    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(clean_text(cell) for cell in row):
            continue
        yield row_number, row


def employee_roster_value(row: Tuple[Any, ...], headers: List[str], field: str) -> str:
    return text_val(row, headers, EMPLOYEE_ROSTER_ALIASES[field])


def employee_header_variants(value: Any) -> set:
    raw = normalize_header(value)
    variants = set()
    for segment in [raw, *re.split(r'[/|\\()\[\]{}]+', raw)]:
        cleaned = re.sub(r'[^\w\u0600-\u06ff]+', ' ', segment.replace('_', ' '))
        cleaned = ' '.join(cleaned.split())
        if cleaned:
            variants.add(cleaned)
    return variants


def employee_header_matches(header: Any, aliases: Iterable[str]) -> bool:
    header_variants = employee_header_variants(header)
    alias_variants = {
        variant
        for alias in aliases
        for variant in employee_header_variants(alias)
    }
    return bool(alias_variants & header_variants)


def find_employee_header_row(ws) -> Tuple[int, List[str]]:
    """Find a roster header by recognized employee fields, not sheet position/name."""
    best_row = 0
    best_headers: List[str] = []
    best_matched_fields = 0
    best_score = 0
    max_row = min(ws.max_row or 30, 30)
    fallback_row = 0
    fallback_headers: List[str] = []
    fallback_non_empty = 0
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, values_only=True),
        start=1,
    ):
        headers = [clean_text(cell) for cell in row]
        non_empty = [cell for cell in headers if clean_text(cell)]
        matched_fields = sum(
            1 for aliases in EMPLOYEE_ROSTER_ALIASES.values()
            if any(employee_header_matches(cell, aliases) for cell in headers)
        )
        score = (matched_fields * 20) + len(non_empty)
        if len(non_empty) > fallback_non_empty:
            fallback_row = row_number
            fallback_headers = headers
            fallback_non_empty = len(non_empty)
        if matched_fields >= 2 and score > best_score:
            best_row = row_number
            best_headers = headers
            best_matched_fields = matched_fields
            best_score = score
    if not best_headers or best_matched_fields < 2:
        if fallback_headers and fallback_non_empty >= 2:
            return fallback_row, fallback_headers
        raise ValueError('Could not detect a usable header row with at least two columns.')
    return best_row, best_headers


def employee_field_candidates(headers: List[str]) -> Dict[str, List[int]]:
    return {
        field: [
            index for index, header in enumerate(headers)
            if clean_text(header) and employee_header_matches(header, aliases)
        ]
        for field, aliases in EMPLOYEE_ROSTER_ALIASES.items()
    }


def resolve_employee_field_mapping(
    headers: List[str],
    requested_mapping: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[str, Optional[int]], Dict[str, List[int]]]:
    candidates = employee_field_candidates(headers)
    mapping: Dict[str, Optional[int]] = {}
    supplied = requested_mapping if isinstance(requested_mapping, dict) else None
    for field in EMPLOYEE_ROSTER_ALIASES:
        if supplied is not None and field in supplied:
            raw_index = supplied.get(field)
            if raw_index in (None, '', -1, '-1'):
                mapping[field] = None
                continue
            try:
                index = int(raw_index)
            except (TypeError, ValueError):
                raise ValueError(f'Invalid column mapping for {field}.')
            if index < 0 or index >= len(headers) or not clean_text(headers[index]):
                raise ValueError(f'Column mapping for {field} is outside the detected header range.')
            mapping[field] = index
        else:
            field_candidates = candidates[field]
            mapping[field] = field_candidates[0] if len(field_candidates) == 1 else None

    used_columns: Dict[int, List[str]] = {}
    for field, index in mapping.items():
        if index is not None:
            used_columns.setdefault(index, []).append(field)
    duplicate_assignments = [fields for fields in used_columns.values() if len(fields) > 1]
    if duplicate_assignments:
        fields = ', '.join('/'.join(items) for items in duplicate_assignments)
        raise ValueError(f'One Excel column cannot be mapped to multiple employee fields: {fields}.')
    return mapping, candidates


def mapped_employee_cell(row: Tuple[Any, ...], mapping: Dict[str, Optional[int]], field: str):
    index = mapping.get(field)
    if index is None or index >= len(row):
        return ''
    return row[index]


def employee_mapping_options(ws, headers: List[str], header_row: int) -> List[Dict[str, Any]]:
    samples: Dict[int, List[str]] = {index: [] for index in range(len(headers))}
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=min(ws.max_row or header_row + 5, header_row + 5),
        values_only=True,
    ):
        for index, value in enumerate(row[:len(headers)]):
            text = clean_text(value)
            if text and text not in samples[index] and len(samples[index]) < 2:
                samples[index].append(text[:80])
    return [
        {'index': index, 'header': clean_text(header), 'samples': samples.get(index, [])}
        for index, header in enumerate(headers)
        if clean_text(header)
    ]


def employee_choice(field: str, value: Any, default: str = '') -> Optional[str]:
    text = normalize_header(value)
    if not text:
        return default
    return EMPLOYEE_CHOICE_MAPS[field].get(text)


def review_fingerprint(file_name: str, sheet_name: str, row_number: int, raw_payload: Dict[str, Any]) -> str:
    material = json.dumps(
        {'file': file_name, 'sheet': sheet_name, 'row': row_number, 'payload': raw_payload},
        ensure_ascii=False,
        sort_keys=True,
        default=str,
    )
    return hashlib.sha256(material.encode('utf-8')).hexdigest()


def import_employee_roster_sheet(
    ws,
    sheet_name: str,
    file_name: str,
    commit: bool,
    user=None,
    flexible: bool = False,
    field_mapping: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Activate valid employees and quarantine invalid/conflicting rows for review."""
    header_row, headers = find_employee_header_row(ws) if flexible else find_header_row(ws)
    column_mapping, mapping_candidates = resolve_employee_field_mapping(headers, field_mapping)
    missing_columns = [field for field in EMPLOYEE_REQUIRED_FIELDS if column_mapping.get(field) is None]
    if missing_columns and not flexible:
        raise ValueError(f"Missing required employee columns: {', '.join(missing_columns)}")

    centers_by_name = {
        normalize_header(center.name): center
        for center in HealthCenter.objects.filter(is_active=True)
    }
    existing_id_map = dict(Employee.objects.exclude(national_id='').values_list('national_id', 'id'))
    existing_email_map = dict(
        Employee.objects.exclude(email__isnull=True).exclude(email='').values_list('email', 'id')
    )
    existing_number_map = dict(
        Employee.objects.exclude(employee_number__isnull=True).exclude(employee_number='')
        .values_list('employee_number', 'id')
    )
    seen_ids = set()
    seen_emails = set()
    seen_numbers = set()
    valid_payloads: List[Dict[str, Any]] = []
    duplicates: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    preview_rows: List[Dict[str, Any]] = []
    review_candidates: List[Dict[str, Any]] = []
    total_rows = 0
    today = date.today()

    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        raw_input = {
            field: clean_text(mapped_employee_cell(row, column_mapping, field))
            for field in EMPLOYEE_ROSTER_ALIASES
        }
        raw = dict(raw_input)
        raw['email'] = raw['email'].strip().lower()
        raw['national_id'] = ''.join(ch for ch in raw['national_id'] if ch.isdigit())
        raw['employee_number'] = raw['employee_number'].strip()
        raw['mobile'] = ''.join(ch for ch in raw['mobile'] if ch.isdigit() or ch == '+')
        row_errors: List[str] = []

        for field in EMPLOYEE_REQUIRED_FIELDS:
            if not raw.get(field):
                row_errors.append(f'Missing required field: {field}.')
        try:
            validate_email(raw['email'])
        except DjangoValidationError:
            row_errors.append('Invalid email address.')
        if raw['national_id'] and len(raw['national_id']) != 10:
            row_errors.append('National ID must contain exactly 10 digits.')
        mobile_digits = ''.join(ch for ch in raw['mobile'] if ch.isdigit())
        if raw['mobile'] and not (9 <= len(mobile_digits) <= 15):
            row_errors.append('Mobile number must contain 9 to 15 digits.')

        date_of_birth = parse_date(mapped_employee_cell(row, column_mapping, 'date_of_birth'))
        appointment_date = parse_date(mapped_employee_cell(row, column_mapping, 'appointment_date'))
        if raw['date_of_birth'] and not date_of_birth:
            row_errors.append('Invalid date of birth. Use a real Excel date or YYYY-MM-DD.')
        if raw['appointment_date'] and not appointment_date:
            row_errors.append('Invalid appointment date. Use a real Excel date or YYYY-MM-DD.')
        if date_of_birth and date_of_birth > today:
            row_errors.append('Date of birth cannot be in the future.')
        if appointment_date and appointment_date > today:
            row_errors.append('Appointment date cannot be in the future.')
        if date_of_birth and appointment_date and appointment_date <= date_of_birth:
            row_errors.append('Appointment date must be after date of birth.')

        gender = employee_choice('gender', raw['gender'])
        marital_status = employee_choice('marital_status', raw['marital_status'])
        periodic_exam_status = employee_choice('periodic_exam_status', raw['periodic_exam_status'], 'incomplete')
        vaccination_status = employee_choice('vaccination_status', raw['vaccination_status'], 'due')
        risk_level = employee_choice('risk_level', raw['risk_level'], 'low')
        if raw['gender'] and not gender:
            row_errors.append('Invalid gender. Use male/female or ذكر/أنثى.')
        if raw['marital_status'] and not marital_status:
            row_errors.append('Invalid marital status.')
        for field, parsed in (
            ('periodic_exam_status', periodic_exam_status),
            ('vaccination_status', vaccination_status),
            ('risk_level', risk_level),
        ):
            if raw[field] and not parsed:
                row_errors.append(f'Invalid value for {field}.')

        center = centers_by_name.get(normalize_header(raw['health_center']))
        if raw['health_center'] and not center:
            row_errors.append('Health center does not match an active center in the platform.')

        duplicate_reasons: List[str] = []
        if raw['national_id'] in seen_ids:
            duplicate_reasons.append('National ID is duplicated inside the file.')
        if raw['email'] in seen_emails:
            duplicate_reasons.append('Email is duplicated inside the file.')
        if raw['employee_number'] in seen_numbers:
            duplicate_reasons.append('Employee number is duplicated inside the file.')
        if raw['national_id'] in existing_id_map:
            duplicate_reasons.append('National ID already exists in PostgreSQL.')
        if raw['email'] in existing_email_map:
            duplicate_reasons.append('Email already exists in PostgreSQL.')
        if raw['employee_number'] in existing_number_map:
            duplicate_reasons.append('Employee number already exists in PostgreSQL.')

        review_payload = {
            'name': raw['name'],
            'email': raw['email'],
            'national_id': raw['national_id'],
            'employee_number': raw['employee_number'],
            'national_address': raw['national_address'],
            'mobile': raw['mobile'],
            'date_of_birth': date_of_birth.isoformat() if date_of_birth else '',
            'birth_place': raw['birth_place'],
            'gender': gender or raw['gender'],
            'marital_status': marital_status or raw['marital_status'],
            'health_center_name': center.name if center else raw['health_center'],
            'health_center': center.id if center else None,
            'job_title': raw['job_title'],
            'appointment_date': appointment_date.isoformat() if appointment_date else '',
            'periodic_exam_status': periodic_exam_status or raw['periodic_exam_status'],
            'vaccination_status': vaccination_status or raw['vaccination_status'],
            'risk_level': risk_level or raw['risk_level'],
        }
        conflict_employee_id = (
            existing_id_map.get(raw['national_id'])
            or existing_email_map.get(raw['email'])
            or existing_number_map.get(raw['employee_number'])
        )

        if row_errors:
            errors.append({'row': row_number, 'reason': ' '.join(row_errors)})
            review_candidates.append({
                'row': row_number,
                'status': 'pending',
                'raw_payload': raw_input,
                'employee_payload': review_payload,
                'issues': row_errors,
                'conflict_employee_id': conflict_employee_id,
            })
            if raw['national_id']:
                seen_ids.add(raw['national_id'])
            if raw['email']:
                seen_emails.add(raw['email'])
            if raw['employee_number']:
                seen_numbers.add(raw['employee_number'])
            continue
        if duplicate_reasons:
            duplicates.append({
                'row': row_number,
                'national_id': mask_value(raw['national_id']),
                'name': raw['name'],
                'reason': ' '.join(duplicate_reasons) + ' The row requires review before activation.',
            })
            review_candidates.append({
                'row': row_number,
                'status': 'conflict',
                'raw_payload': raw_input,
                'employee_payload': review_payload,
                'issues': duplicate_reasons,
                'conflict_employee_id': conflict_employee_id,
            })
            if raw['national_id']:
                seen_ids.add(raw['national_id'])
            if raw['email']:
                seen_emails.add(raw['email'])
            if raw['employee_number']:
                seen_numbers.add(raw['employee_number'])
            continue

        payload = {
            'name': raw['name'],
            'email': raw['email'],
            'national_id': raw['national_id'],
            'employee_number': raw['employee_number'],
            'national_address': raw['national_address'],
            'mobile': raw['mobile'],
            'date_of_birth': date_of_birth,
            'birth_place': raw['birth_place'],
            'gender': gender,
            'marital_status': marital_status,
            'health_center': center.name if center else '',
            'job_title': raw['job_title'],
            'appointment_date': appointment_date,
            'periodic_exam_status': periodic_exam_status,
            'vaccination_status': vaccination_status,
            'risk_level': risk_level,
            '_center': center,
        }
        valid_payloads.append(payload)
        seen_ids.add(raw['national_id'])
        seen_emails.add(raw['email'])
        seen_numbers.add(raw['employee_number'])
        if len(preview_rows) < 20:
            preview_rows.append({'row': row_number, **masked_preview(payload)})

    summary = {
        'total_rows': total_rows,
        'valid_rows': len(valid_payloads),
        'duplicate_rows': len(duplicates),
        'errors_count': len(errors),
        'skipped_rows': 0,
        'imported_employees': 0,
        'pending_reviews': len([item for item in review_candidates if item['status'] == 'pending']),
        'conflict_reviews': len([item for item in review_candidates if item['status'] == 'conflict']),
        'staged_for_review': len(review_candidates),
    }
    review_ids: List[int] = []
    if commit:
        with transaction.atomic():
            for payload in valid_payloads:
                employee_values = {
                    key: value for key, value in payload.items()
                    if key not in {'health_center', '_center'}
                }
                Employee.objects.create(health_center=payload['_center'], **employee_values)
                summary['imported_employees'] += 1
            for candidate in review_candidates:
                fingerprint = review_fingerprint(
                    file_name,
                    sheet_name,
                    candidate['row'],
                    candidate['raw_payload'],
                )
                review, created = EmployeeImportReview.objects.get_or_create(
                    fingerprint=fingerprint,
                    defaults={
                        'source_file': file_name,
                        'source_sheet': sheet_name,
                        'source_row': candidate['row'],
                        'raw_payload': candidate['raw_payload'],
                        'employee_payload': candidate['employee_payload'],
                        'issues': candidate['issues'],
                        'status': candidate['status'],
                        'conflict_employee_id': candidate['conflict_employee_id'],
                        'created_by': user,
                    },
                )
                if not created and review.status in ('pending', 'conflict'):
                    review.raw_payload = candidate['raw_payload']
                    review.employee_payload = candidate['employee_payload']
                    review.issues = candidate['issues']
                    review.status = candidate['status']
                    review.conflict_employee_id = candidate['conflict_employee_id']
                    review.save(update_fields=[
                        'raw_payload', 'employee_payload', 'issues', 'status',
                        'conflict_employee', 'updated_at',
                    ])
                review_ids.append(review.id)

    used_indices = {index for index in column_mapping.values() if index is not None}
    return {
        'detected_headers': headers,
        'mapped_fields': column_mapping,
        'mapping_options': employee_mapping_options(ws, headers, header_row),
        'mapping_candidates': mapping_candidates,
        'ignored_columns': [
            {'index': index, 'header': clean_text(header)}
            for index, header in enumerate(headers)
            if clean_text(header) and index not in used_indices
        ],
        'missing_required_fields': missing_columns,
        'header_row': header_row,
        'employee_import_mode': 'flexible' if flexible else 'template',
        'invalid_policy': 'stage_for_review',
        'summary': summary,
        'duplicates': duplicates[:100],
        'errors': errors[:100],
        'preview_rows': preview_rows,
        'review_ids': review_ids,
    }


def database_payload(row: Tuple[Any, ...], headers: List[str], row_number: int, sheet_name: str, file_name: str) -> Dict[str, Any]:
    return {
        'name': text_val(row, headers, ['Name']),
        'national_id': text_val(row, headers, ['National ID']),
        'mobile': text_val(row, headers, ['Mobile #', 'Mobile']),
        'gender': parse_gender(text_val(row, headers, ['Gender'])),
        'health_center': text_val(row, headers, ['Health center']),
        'job_title': text_val(row, headers, ['Job Title']),
        'age': parse_int(text_val(row, headers, ['Age'])),
        'periodic_exam_status': 'completed' if normalize_header(text_val(row, headers, ['Latest Field Visit', 'Latest Virtual Visit'])) else 'incomplete',
        'vaccination_status': 'completed' if normalize_header(text_val(row, headers, ['FLU Vaccine', 'HBV Vaccine'])) in ('yes', 'completed', 'done') else 'due',
        'risk_level': 'medium' if normalize_header(text_val(row, headers, ['Any Chronic Disease', 'Restrictions'])) in ('yes', 'true') else 'low',
        '_trace': make_trace(headers, row, row_number, sheet_name, file_name),
        'marital_status': text_val(row, headers, ['Marital Status']),
        'children_count': parse_int(text_val(row, headers, ['# of Child']), 0) if text_val(row, headers, ['# of Child']) else None,
        'moh_id': text_val(row, headers, ['MOH ID']),
        'employment_start_date': parse_date(val(row, headers, ['Date of Start (Old)'])),
        'experience_years': parse_decimal(text_val(row, headers, ['Years of Experience'])),
        'current_position': text_val(row, headers, ['Current Job']),
        'weight_kg': parse_decimal(text_val(row, headers, ['Wt (Kg)'])),
        'height_cm': parse_decimal(text_val(row, headers, ['Ht (cm)'])),
        'bmi': parse_decimal(text_val(row, headers, ['BMI'])),
        'obesity_status': text_val(row, headers, ['Obesity']),
        'physical_activity': text_val(row, headers, ['Physical Exercise']),
        'diabetes': parse_bool(text_val(row, headers, ['DM (Diabetes Mellitus)'])),
        'hypertension': parse_bool(text_val(row, headers, ['HTN (Hypertension)'])),
        'thyroid_disease': parse_bool(text_val(row, headers, ['Thyroid'])),
        'asthma': parse_bool(text_val(row, headers, ['Asthma'])),
        'blood_disease': parse_bool(text_val(row, headers, ['Blood Disorder'])),
        'smoking_status': text_val(row, headers, ['Smoking']),
        'surgical_history': text_val(row, headers, ['Surgical Hx(if any)', 'Surgical Hx']),
        'family_history': text_val(row, headers, ['Other + F.Hx']),
        'medical_restrictions': text_val(row, headers, ['Restrictions']),
        'notes': text_val(row, headers, ['Comments']),
    }


def import_database_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    valid_payloads: List[Dict[str, Any]] = []
    duplicates: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    preview_rows: List[Dict[str, Any]] = []
    seen_ids = set()
    existing_ids = set(Employee.objects.values_list('national_id', flat=True))
    total_rows = 0

    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        payload = database_payload(row, headers, row_number, sheet_name, file_name)
        national_id = payload['national_id']
        if not payload['name'] or not national_id:
            errors.append({'row': row_number, 'reason': 'Missing employee name or national ID.'})
            continue
        duplicate_reason = ''
        if national_id in seen_ids:
            duplicate_reason = 'Duplicate national ID inside the uploaded sheet.'
        elif national_id in existing_ids:
            duplicate_reason = 'National ID already exists in PostgreSQL; record will be updated in commit mode.'
        if duplicate_reason and not commit:
            duplicates.append({'row': row_number, 'national_id': mask_value(national_id), 'name': payload['name'], 'reason': duplicate_reason})
        seen_ids.add(national_id)
        valid_payloads.append(payload)
        if len(preview_rows) < 12:
            preview_rows.append({'row': row_number, **masked_preview(payload)})

    summary = {'total_rows': total_rows, 'valid_rows': len(valid_payloads), 'duplicate_rows': len(duplicates), 'errors_count': len(errors), 'skipped_rows': len(errors) + (len(duplicates) if not commit else 0), 'imported_employees': 0, 'imported_health_profiles': 0, 'imported_lab_screenings': 0, 'imported_vaccination_records': 0, 'imported_screening_records': 0}

    if commit:
        with transaction.atomic():
            for payload in valid_payloads:
                employee = upsert_employee(payload)
                trace = payload['_trace']
                raw = trace['raw_payload']
                EmployeeHealthProfile.objects.update_or_create(
                    employee=employee,
                    defaults={**trace, 'marital_status': payload['marital_status'], 'children_count': payload['children_count'], 'moh_id': payload['moh_id'], 'employment_start_date': payload['employment_start_date'], 'experience_years': payload['experience_years'], 'current_position': payload['current_position'], 'weight_kg': payload['weight_kg'], 'height_cm': payload['height_cm'], 'bmi': payload['bmi'], 'obesity_status': payload['obesity_status'], 'physical_activity': payload['physical_activity'], 'diabetes': payload['diabetes'], 'hypertension': payload['hypertension'], 'thyroid_disease': payload['thyroid_disease'], 'asthma': payload['asthma'], 'blood_disease': payload['blood_disease'], 'smoking_status': payload['smoking_status'], 'surgical_history': payload['surgical_history'], 'family_history': payload['family_history'], 'medical_restrictions': payload['medical_restrictions'], 'notes': payload['notes']},
                )
                summary['imported_employees'] += 1
                summary['imported_health_profiles'] += 1

                if clean_text(raw.get('LAB Request')):
                    LabScreening.objects.update_or_create(
                        employee=employee,
                        request_date=parse_date(raw.get('LAB Request Date')),
                        defaults={**trace, 'request_status': 'completed' if normalize_header(raw.get('LAB Result')) in ('yes', 'completed') else 'requested', 'result_status': clean_text(raw.get('LAB Result')), 'result_checked_date': parse_date(raw.get('Lab Result Checked Date')), 'anti_hbs': clean_text(raw.get('Anti-HBs')), 'hbsag': clean_text(raw.get('HBsAg')), 'hcv': clean_text(raw.get('HCV (Hepatitis C Virus)')), 'hiv': clean_text(raw.get('HIV')), 'rubella_igg': clean_text(raw.get('Rubella IGg')), 'measles_igg': clean_text(raw.get('Measles IGg')), 'varicella_igg': clean_text(raw.get('Varicella IGg')), 'ppd_test': clean_text(raw.get('PPD Test')), 'follow_up_required': bool(clean_text(raw.get('Follow-up')))},
                    )
                    summary['imported_lab_screenings'] += 1

                vaccines = [('HBV', 'HBV Vaccine', 'HBV Vac. 1st Dose', 'HBV Vac. 2nd Dose', 'HBV Vac. 3rd Dose', 'Post Vac. Anti-HBs'), ('MMR', 'MMR Vaccine', 'MMR Vac. 1st Dose', 'MMR Vac. 2nd Dose', None, None), ('FLU', 'FLU Vaccine', None, None, None, None), ('MCV4', 'MCV4 Vaccine', None, None, None, None), ('COVID-19', 'COVID-19 Vaccination', None, None, None, None)]
                for vaccine_type, status_col, d1, d2, d3, post_col in vaccines:
                    status_text = clean_text(raw.get(status_col))
                    if not status_text:
                        continue
                    VaccinationRecord.objects.update_or_create(
                        employee=employee,
                        vaccine_type=vaccine_type,
                        campaign_name='',
                        defaults={**trace, 'status': 'completed' if normalize_header(status_text) in ('yes', 'given', 'done', 'completed') else 'due', 'first_dose_date': parse_date(raw.get(d1)) if d1 else None, 'second_dose_date': parse_date(raw.get(d2)) if d2 else None, 'third_dose_date': parse_date(raw.get(d3)) if d3 else None, 'post_vaccine_result': clean_text(raw.get(post_col)) if post_col else '', 'notes': status_text},
                    )
                    summary['imported_vaccination_records'] += 1

                for program, result_col in (('colon_cancer', 'Colon Cancer Screening'), ('other', 'Breast Cancer Screening')):
                    result = clean_text(raw.get(result_col))
                    if not result:
                        continue
                    ScreeningProgramRecord.objects.update_or_create(employee=employee, program_type=program, program_year=2025, defaults={**trace, 'status': 'completed' if normalize_header(result) in ('yes', 'done', 'completed') else 'pending', 'result': result})
                    summary['imported_screening_records'] += 1

    return {'detected_headers': headers, 'mapped_fields': {'mode': 'database_full_profile'}, 'summary': summary, 'duplicates': duplicates[:50], 'errors': errors[:50], 'preview_rows': preview_rows}


def import_employee_clinic_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    payloads = []
    errors = []
    preview_rows = []
    total_rows = 0
    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        payload = {'name': text_val(row, headers, HEADER_ALIASES['name']), 'national_id': text_val(row, headers, HEADER_ALIASES['national_id']), 'mobile': text_val(row, headers, HEADER_ALIASES['mobile']), 'job_title': text_val(row, headers, HEADER_ALIASES['job_title']), 'health_center': text_val(row, headers, HEADER_ALIASES['health_center']), 'visit_date': parse_date(val(row, headers, ['التاريخ', 'م'])), 'weight_kg': parse_decimal(text_val(row, headers, ['الوزن'])), 'diagnosis': text_val(row, headers, ['التشخيص']), 'sick_leave_days': parse_int(''.join(ch for ch in text_val(row, headers, ['عدد الايام']) if ch.isdigit()), 0) or None, 'physician': text_val(row, headers, ['الطبيب']), 'follow_up_date': parse_date(val(row, headers, ['موعد المتابعة الافتراضي'])), '_trace': make_trace(headers, row, row_number, sheet_name, file_name)}
        if not payload['name'] or not payload['national_id']:
            errors.append({'row': row_number, 'reason': 'Missing employee name or national ID.'})
            continue
        payloads.append(payload)
        if len(preview_rows) < 12:
            preview_rows.append({'row': row_number, **masked_preview(payload)})
    summary = {'total_rows': total_rows, 'valid_rows': len(payloads), 'duplicate_rows': 0, 'errors_count': len(errors), 'skipped_rows': len(errors), 'imported_employee_clinic_visits': 0, 'imported_employees': 0}
    if commit:
        with transaction.atomic():
            for payload in payloads:
                employee = upsert_employee({**payload, 'gender': 'male', 'age': 0})
                update_or_create_by_filter(EmployeeClinicVisit, {'employee': employee, 'visit_date': payload['visit_date'], 'diagnosis': payload['diagnosis']}, {**payload['_trace'], 'weight_kg': payload['weight_kg'], 'sick_leave_days': payload['sick_leave_days'], 'physician': payload['physician'], 'follow_up_date': payload['follow_up_date'], 'notes': ''})
                summary['imported_employees'] += 1
                summary['imported_employee_clinic_visits'] += 1
    return {'detected_headers': headers, 'mapped_fields': {'mode': 'employee_clinic'}, 'summary': summary, 'duplicates': [], 'errors': errors[:50], 'preview_rows': preview_rows}


def import_occupational_clinic_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    payloads, errors, preview_rows = [], [], []
    total_rows = 0
    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        payload = {'name': text_val(row, headers, HEADER_ALIASES['name']), 'national_id': text_val(row, headers, HEADER_ALIASES['national_id']), 'mobile': text_val(row, headers, HEADER_ALIASES['mobile']), 'job_title': text_val(row, headers, HEADER_ALIASES['job_title']), 'health_center': text_val(row, headers, HEADER_ALIASES['health_center']), 'contract_type': text_val(row, headers, ['نوع العقد']), 'action': text_val(row, headers, ['الاجراء', 'الإجراء']), 'physician': text_val(row, headers, ['الطبيب']), 'visit_date': parse_date(val(row, headers, ['م', 'التاريخ'])), '_trace': make_trace(headers, row, row_number, sheet_name, file_name)}
        if not payload['name'] or not payload['national_id']:
            errors.append({'row': row_number, 'reason': 'Missing employee name or national ID.'})
            continue
        payloads.append(payload)
        if len(preview_rows) < 12:
            preview_rows.append({'row': row_number, **masked_preview(payload)})
    summary = {'total_rows': total_rows, 'valid_rows': len(payloads), 'duplicate_rows': 0, 'errors_count': len(errors), 'skipped_rows': len(errors), 'imported_occupational_clinic_visits': 0, 'imported_employees': 0}
    if commit:
        with transaction.atomic():
            for payload in payloads:
                employee = upsert_employee({**payload, 'gender': 'male', 'age': 0})
                update_or_create_by_filter(OccupationalClinicVisit, {'employee': employee, 'visit_date': payload['visit_date'], 'action': payload['action']}, {**payload['_trace'], 'contract_type': payload['contract_type'], 'physician': payload['physician'], 'notes': ''})
                summary['imported_employees'] += 1
                summary['imported_occupational_clinic_visits'] += 1
    return {'detected_headers': headers, 'mapped_fields': {'mode': 'occupational_clinic'}, 'summary': summary, 'duplicates': [], 'errors': errors[:50], 'preview_rows': preview_rows}


def import_medical_committee_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    payloads, errors, preview_rows = [], [], []
    total_rows = 0
    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        payload = {'name': text_val(row, headers, HEADER_ALIASES['name']), 'national_id': text_val(row, headers, HEADER_ALIASES['national_id']), 'mobile': text_val(row, headers, HEADER_ALIASES['mobile']), 'job_title': text_val(row, headers, HEADER_ALIASES['job_title']), 'health_center': text_val(row, headers, HEADER_ALIASES['health_center']), 'contract_type': text_val(row, headers, ['نوع العقد']), 'transaction_number': text_val(row, headers, ['رقم المعاملة', 'UCC']), 'diagnosis': text_val(row, headers, ['التشخيص']), 'recommendations': text_val(row, headers, ['Recommendations', 'التوصيات', 'قرار الصلاحية']), 'decision_date': parse_date(val(row, headers, ['تاريخ القرار'])), 'physician': text_val(row, headers, ['الطبيب']), '_trace': make_trace(headers, row, row_number, sheet_name, file_name)}
        if not payload['name'] or not payload['national_id']:
            errors.append({'row': row_number, 'reason': 'Missing employee name or national ID.'})
            continue
        payloads.append(payload)
        if len(preview_rows) < 12:
            preview_rows.append({'row': row_number, **masked_preview(payload)})
    summary = {'total_rows': total_rows, 'valid_rows': len(payloads), 'duplicate_rows': 0, 'errors_count': len(errors), 'skipped_rows': len(errors), 'imported_medical_committee_cases': 0, 'imported_employees': 0}
    if commit:
        with transaction.atomic():
            for payload in payloads:
                employee = upsert_employee({**payload, 'gender': 'male', 'age': 0})
                filter_kwargs = {'transaction_number': payload['transaction_number']} if payload['transaction_number'] else {'employee': employee, 'decision_date': payload['decision_date'], 'diagnosis': payload['diagnosis']}
                update_or_create_by_filter(MedicalCommitteeCase, filter_kwargs, {**payload['_trace'], 'employee': employee, 'contract_type': payload['contract_type'], 'diagnosis': payload['diagnosis'], 'recommendations': payload['recommendations'], 'decision': payload['recommendations'], 'decision_date': payload['decision_date'], 'physician': payload['physician'], 'status': 'decision_issued' if payload['decision_date'] else 'under_review'})
                summary['imported_employees'] += 1
                summary['imported_medical_committee_cases'] += 1
    return {'detected_headers': headers, 'mapped_fields': {'mode': 'medical_committee'}, 'summary': summary, 'duplicates': [], 'errors': errors[:50], 'preview_rows': preview_rows}


def import_needle_stick_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    payloads, errors, preview_rows = [], [], []
    total_rows = 0
    source_col = col_index(headers, ['المصدر', 'source'])
    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        payload = {'name': text_val(row, headers, HEADER_ALIASES['name']), 'national_id': text_val(row, headers, HEADER_ALIASES['national_id']), 'health_center': text_val(row, headers, ['مكان العمل', 'المركز']), 'job_title': text_val(row, headers, ['طبيعة العمل', 'مسمى الوظيفة']), 'exposure_date': parse_date(val(row, headers, ['تاريخ التعرض'])), 'work_nature': text_val(row, headers, ['طبيعة العمل']), 'workplace': text_val(row, headers, ['مكان العمل']), 'anti_hbs_result': text_val(row, headers, ['نتيجة Anti-HBs', 'Anti-HBs']), 'injury_method': text_val(row, headers, ['طريقة الاصابة', 'طريقة الإصابة']), 'source_known': parse_bool(text_val(row, headers, ['المصدر'])), 'source_name': clean_text(row[source_col + 1]) if source_col is not None and source_col + 1 < len(row) else '', 'source_national_id': clean_text(row[source_col + 2]) if source_col is not None and source_col + 2 < len(row) else '', 'source_result': text_val(row, headers, ['نتيجة المصدر']), 'referred_to_employee_clinic': parse_bool(text_val(row, headers, ['التحويل الى عيادة الموظفين في البرج'])) or False, '_trace': make_trace(headers, row, row_number, sheet_name, file_name)}
        if not payload['name'] or not payload['national_id']:
            errors.append({'row': row_number, 'reason': 'Missing employee name or national ID.'})
            continue
        payloads.append(payload)
        if len(preview_rows) < 12:
            preview_rows.append({'row': row_number, **masked_preview(payload)})
    summary = {'total_rows': total_rows, 'valid_rows': len(payloads), 'duplicate_rows': 0, 'errors_count': len(errors), 'skipped_rows': len(errors), 'imported_needle_stick_exposures': 0, 'imported_employees': 0}
    if commit:
        with transaction.atomic():
            for payload in payloads:
                employee = upsert_employee({**payload, 'mobile': '', 'gender': 'male', 'age': 0})
                update_or_create_by_filter(NeedleStickExposure, {'employee': employee, 'exposure_date': payload['exposure_date'], 'injury_method': payload['injury_method']}, {**payload['_trace'], 'work_nature': payload['work_nature'], 'workplace': payload['workplace'], 'anti_hbs_result': payload['anti_hbs_result'], 'source_known': payload['source_known'], 'source_name': payload['source_name'], 'source_national_id': payload['source_national_id'], 'source_result': payload['source_result'], 'referred_to_employee_clinic': payload['referred_to_employee_clinic'], 'status': 'referred' if payload['referred_to_employee_clinic'] else 'new'})
                summary['imported_employees'] += 1
                summary['imported_needle_stick_exposures'] += 1
    return {'detected_headers': headers, 'mapped_fields': {'mode': 'needle_stick'}, 'summary': summary, 'duplicates': [], 'errors': errors[:50], 'preview_rows': preview_rows}


def import_vaccination_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    payloads, errors, preview_rows = [], [], []
    total_rows = 0
    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        payload = {'name': text_val(row, headers, HEADER_ALIASES['name']), 'national_id': text_val(row, headers, HEADER_ALIASES['national_id']), 'health_center': text_val(row, headers, HEADER_ALIASES['health_center']), 'discovery_date': parse_date(val(row, headers, ['تاريخ الأكتشاف', 'تاريخ الاكتشاف'])), 'vaccine_type': text_val(row, headers, ['نوع اللقاح']) or 'HBV', 'first_dose_date': parse_date(val(row, headers, ['الأولى', 'الاولى'])), 'second_dose_date': parse_date(val(row, headers, ['الثانية'])), 'third_dose_date': parse_date(val(row, headers, ['الثالثة'])), 'notes': text_val(row, headers, ['ملاحظات']), 'follow_up': text_val(row, headers, ['مواعيد افتراضية']), '_trace': make_trace(headers, row, row_number, sheet_name, file_name)}
        if not payload['name']:
            errors.append({'row': row_number, 'reason': 'Missing employee name.'})
            continue
        if not payload['national_id']:
            payload['national_id'] = f"NO-ID-{sheet_name}-{row_number}"
        payloads.append(payload)
        if len(preview_rows) < 12:
            preview_rows.append({'row': row_number, **masked_preview(payload)})
    summary = {'total_rows': total_rows, 'valid_rows': len(payloads), 'duplicate_rows': 0, 'errors_count': len(errors), 'skipped_rows': len(errors), 'imported_vaccination_records': 0, 'imported_employees': 0}
    if commit:
        with transaction.atomic():
            for payload in payloads:
                employee = upsert_employee({**payload, 'mobile': '', 'gender': 'male', 'job_title': '', 'age': 0})
                VaccinationRecord.objects.update_or_create(employee=employee, vaccine_type=payload['vaccine_type'], campaign_name='', defaults={**payload['_trace'], 'status': 'completed' if payload['third_dose_date'] or normalize_header(payload['notes']) == 'done' else 'due', 'first_dose_date': payload['first_dose_date'], 'second_dose_date': payload['second_dose_date'], 'third_dose_date': payload['third_dose_date'], 'notes': payload['notes'] or payload['follow_up']})
                summary['imported_employees'] += 1
                summary['imported_vaccination_records'] += 1
    return {'detected_headers': headers, 'mapped_fields': {'mode': 'vaccinations'}, 'summary': summary, 'duplicates': [], 'errors': errors[:50], 'preview_rows': preview_rows}


def import_campaign_coverage_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    payloads, errors, preview_rows = [], [], []
    total_rows = 0
    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        center_name = text_val(row, headers, ['المركز', 'center'])
        if not center_name:
            errors.append({'row': row_number, 'reason': 'Missing health center.'})
            continue
        target = parse_int(text_val(row, headers, ['عدد المستهدفين']), 0)
        completed = parse_int(text_val(row, headers, ['الكشف الدوري', 'التحاليل المكتملة', 'المجموع']), 0)
        payload = {'campaign_name': 'Influenza 2025' if 'انفلونزا' in sheet_name else ('Center Coverage 2025' if 'تغطية' in sheet_name else sheet_name), 'health_center': center_name, 'year': 2025, 'target_count': target, 'completed_count': completed, 'refused_count': parse_int(text_val(row, headers, ['عدد الرفض', 'رفض PPD']), 0), 'contraindicated_count': parse_int(text_val(row, headers, ['عدد الموانع']), 0), 'coverage_percent': parse_decimal(text_val(row, headers, ['%', '% الكشف الدوري', '% من قام بعمل التحاليل'])), 'notes': '', '_trace': make_trace(headers, row, row_number, sheet_name, file_name)}
        payloads.append(payload)
        if len(preview_rows) < 12:
            preview_rows.append({'row': row_number, **masked_preview(payload)})
    summary = {'total_rows': total_rows, 'valid_rows': len(payloads), 'duplicate_rows': 0, 'errors_count': len(errors), 'skipped_rows': len(errors), 'imported_campaign_coverages': 0}
    if commit:
        with transaction.atomic():
            for payload in payloads:
                center = get_or_create_center(payload['health_center'])
                CampaignCoverage.objects.update_or_create(campaign_name=payload['campaign_name'], health_center=center, year=payload['year'], defaults={**payload['_trace'], 'target_count': payload['target_count'], 'completed_count': payload['completed_count'], 'refused_count': payload['refused_count'], 'contraindicated_count': payload['contraindicated_count'], 'coverage_percent': payload['coverage_percent'], 'notes': payload['notes']})
                summary['imported_campaign_coverages'] += 1
    return {'detected_headers': headers, 'mapped_fields': {'mode': 'campaign_coverage'}, 'summary': summary, 'duplicates': [], 'errors': errors[:50], 'preview_rows': preview_rows}


def import_colon_cancer_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    payloads, errors, preview_rows = [], [], []
    total_rows = 0
    current_center = ''
    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        center_text = text_val(row, headers, ['center', 'المركز'])
        if center_text:
            current_center = center_text
        payload = {'health_center': current_center, 'name': text_val(row, headers, ['Name', 'الأسم', 'الاسم']), 'national_id': text_val(row, headers, ['National ID', 'السجل المدني', 'رقم الهوية']), 'dob': parse_date(val(row, headers, ['DOB'])), 'mobile': text_val(row, headers, ['Mobile #', 'رقم الهاتف']), '_trace': make_trace(headers, row, row_number, sheet_name, file_name)}
        if not payload['name'] or not payload['national_id']:
            errors.append({'row': row_number, 'reason': 'Missing employee name or national ID.'})
            continue
        payloads.append(payload)
        if len(preview_rows) < 12:
            preview_rows.append({'row': row_number, **masked_preview(payload)})
    summary = {'total_rows': total_rows, 'valid_rows': len(payloads), 'duplicate_rows': 0, 'errors_count': len(errors), 'skipped_rows': len(errors), 'imported_screening_records': 0, 'imported_employees': 0}
    if commit:
        with transaction.atomic():
            for payload in payloads:
                employee = upsert_employee({**payload, 'gender': 'male', 'job_title': '', 'age': 0})
                ScreeningProgramRecord.objects.update_or_create(employee=employee, program_type='colon_cancer', program_year=2025, defaults={**payload['_trace'], 'status': 'eligible', 'result': '', 'notes': ''})
                summary['imported_employees'] += 1
                summary['imported_screening_records'] += 1
    return {'detected_headers': headers, 'mapped_fields': {'mode': 'colon_cancer'}, 'summary': summary, 'duplicates': [], 'errors': errors[:50], 'preview_rows': preview_rows}


def import_trainee_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    payloads, errors, preview_rows = [], [], []
    total_rows = 0
    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        payload = {'name': text_val(row, headers, ['الأسم', 'الاسم', 'Name']), 'specialty': text_val(row, headers, ['التخصص']), 'rotation_start_date': parse_date(val(row, headers, ['التاريخ'])), 'duration': text_val(row, headers, ['المدة']), 'health_center': text_val(row, headers, ['المركز']), '_trace': make_trace(headers, row, row_number, sheet_name, file_name)}
        if not payload['name']:
            errors.append({'row': row_number, 'reason': 'Missing trainee name.'})
            continue
        payloads.append(payload)
        if len(preview_rows) < 12:
            preview_rows.append({'row': row_number, **masked_preview(payload)})
    summary = {'total_rows': total_rows, 'valid_rows': len(payloads), 'duplicate_rows': 0, 'errors_count': len(errors), 'skipped_rows': len(errors), 'imported_trainee_rotations': 0}
    if commit:
        with transaction.atomic():
            for payload in payloads:
                center = get_or_create_center(payload['health_center']) if payload['health_center'] else None
                update_or_create_by_filter(TraineeRotation, {'name': payload['name'], 'rotation_start_date': payload['rotation_start_date'], 'health_center': center}, {**payload['_trace'], 'specialty': payload['specialty'], 'duration': payload['duration'], 'supervisor': '', 'notes': ''})
                summary['imported_trainee_rotations'] += 1
    return {'detected_headers': headers, 'mapped_fields': {'mode': 'trainees'}, 'summary': summary, 'duplicates': [], 'errors': errors[:50], 'preview_rows': preview_rows}


def import_ddl_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    payloads, preview_rows = [], []
    category_map = {'health center': 'center', 'job title': 'job_title', 'current job': 'job_title', 'الممارس': 'other', 'anti-hbs': 'lab_test', 'hbsag': 'lab_test', 'hcv': 'lab_test', 'hiv': 'lab_test', 'measles igg': 'lab_test', 'rubella igg': 'lab_test'}
    for row_number, row in iter_data_rows(ws, header_row):
        for idx, cell in enumerate(row):
            label = clean_text(cell)
            if not label or idx >= len(headers):
                continue
            header = clean_text(headers[idx])
            category = category_map.get(normalize_header(header), 'other')
            payload = {'category': category, 'code': '', 'label_ar': label, 'label_en': '', '_trace': make_trace(headers, row, row_number, sheet_name, file_name)}
            payloads.append(payload)
            if len(preview_rows) < 12:
                preview_rows.append({'row': row_number, 'category': category, 'label': label, 'source_column': header})
    summary = {'total_rows': len(payloads), 'valid_rows': len(payloads), 'duplicate_rows': 0, 'errors_count': 0, 'skipped_rows': 0, 'imported_reference_lookups': 0}
    if commit:
        with transaction.atomic():
            for payload in payloads:
                ReferenceLookup.objects.update_or_create(category=payload['category'], code=payload['code'], label_ar=payload['label_ar'], defaults={**payload['_trace'], 'label_en': payload['label_en'], 'is_active': True, 'notes': ''})
                summary['imported_reference_lookups'] += 1
    return {'detected_headers': headers, 'mapped_fields': {'mode': 'ddl_reference'}, 'summary': summary, 'duplicates': [], 'errors': [], 'preview_rows': preview_rows}


def import_generic_sheet(ws, sheet_name: str, file_name: str, commit: bool) -> Dict[str, Any]:
    header_row, headers = find_header_row(ws)
    header_map = {key: col_index(headers, aliases) for key, aliases in HEADER_ALIASES.items()}
    missing = [field for field in ['name', 'national_id'] if header_map.get(field) is None]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    payloads, duplicates, errors, preview_rows = [], [], [], []
    seen_ids = set()
    existing_ids = set(Employee.objects.values_list('national_id', flat=True))
    total_rows = 0
    for row_number, row in iter_data_rows(ws, header_row):
        total_rows += 1
        payload = {'name': text_val(row, headers, HEADER_ALIASES['name']), 'national_id': text_val(row, headers, HEADER_ALIASES['national_id']), 'mobile': text_val(row, headers, HEADER_ALIASES['mobile']), 'gender': parse_gender(text_val(row, headers, HEADER_ALIASES['gender'])), 'health_center': text_val(row, headers, HEADER_ALIASES['health_center']) or DEFAULT_CENTER_NAME, 'job_title': text_val(row, headers, HEADER_ALIASES['job_title']), 'age': parse_int(text_val(row, headers, HEADER_ALIASES['age'])), 'diagnosis': text_val(row, headers, HEADER_ALIASES['diagnosis']), 'clinic_type': text_val(row, headers, HEADER_ALIASES['clinic_type']) or 'Imported Excel Visit', 'action_taken': text_val(row, headers, HEADER_ALIASES['action_taken']), 'visit_date': parse_date(val(row, headers, HEADER_ALIASES['visit_date'])), '_trace': make_trace(headers, row, row_number, sheet_name, file_name)}
        if not payload['name'] or not payload['national_id']:
            errors.append({'row': row_number, 'reason': 'Missing employee name or national ID.'})
            continue
        duplicate_reason = ''
        if payload['national_id'] in seen_ids:
            duplicate_reason = 'Duplicate national ID inside the uploaded sheet.'
        elif payload['national_id'] in existing_ids:
            duplicate_reason = 'National ID already exists in PostgreSQL; record will be updated in commit mode.'
        if duplicate_reason and not commit:
            duplicates.append({'row': row_number, 'national_id': mask_value(payload['national_id']), 'name': payload['name'], 'reason': duplicate_reason})
        seen_ids.add(payload['national_id'])
        payloads.append(payload)
        if len(preview_rows) < 12:
            preview_rows.append({'row': row_number, **masked_preview(payload)})
    summary = {'total_rows': total_rows, 'valid_rows': len(payloads), 'duplicate_rows': len(duplicates), 'errors_count': len(errors), 'skipped_rows': len(errors) + len(duplicates), 'imported_employees': 0, 'imported_clinic_visits': 0}
    if commit:
        with transaction.atomic():
            for payload in payloads:
                employee = upsert_employee(payload)
                summary['imported_employees'] += 1
                if payload.get('diagnosis'):
                    update_or_create_by_filter(ClinicVisit, {'employee': employee, 'visit_date': payload['visit_date'], 'diagnosis': payload['diagnosis']}, {'clinic_type': payload['clinic_type'], 'action_taken': payload['action_taken']})
                    summary['imported_clinic_visits'] += 1
    return {'detected_headers': headers, 'mapped_fields': {'mode': 'generic_employee'}, 'summary': summary, 'duplicates': duplicates[:50], 'errors': errors[:50], 'preview_rows': preview_rows}


SHEET_ROUTERS = (
    ('Employees', import_employee_roster_sheet),
    ('الموظفون', import_employee_roster_sheet),
    ('Database', import_database_sheet),
    ('عيادة الموظفين', import_employee_clinic_sheet),
    ('عيادة الصحة المهنية', import_occupational_clinic_sheet),
    ('الهيئة الطبية', import_medical_committee_sheet),
    ('قرارات الهيئة الاعوام السابقة', import_medical_committee_sheet),
    ('الوخزبالأبر', import_needle_stick_sheet),
    ('الوخز', import_needle_stick_sheet),
    ('تطعيمات', import_vaccination_sheet),
    ('نسبة تغطية', import_campaign_coverage_sheet),
    ('حملة الانفلونزا', import_campaign_coverage_sheet),
    ('colon cancer', import_colon_cancer_sheet),
    ('متدربين', import_trainee_sheet),
    ('DDL', import_ddl_sheet),
)


def resolve_processor(sheet_name: str):
    normalized_sheet = normalize_header(sheet_name)
    for token, processor in SHEET_ROUTERS:
        if normalize_header(token) in normalized_sheet:
            return processor
    return import_generic_sheet


def process_single_sheet(
    workbook,
    selected_sheet: str,
    commit: bool,
    file_name: str,
    user=None,
    force_employee_roster: bool = False,
    employee_import_mode: str = 'template',
    field_mapping: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    ws = workbook[selected_sheet]
    processor = import_employee_roster_sheet if force_employee_roster else resolve_processor(selected_sheet)
    if processor is import_employee_roster_sheet:
        result = processor(
            ws,
            selected_sheet,
            file_name,
            commit,
            user=user,
            flexible=employee_import_mode == 'flexible',
            field_mapping=field_mapping,
        )
    else:
        result = processor(ws, selected_sheet, file_name, commit)
    result.update({'mode': 'commit' if commit else 'preview', 'file_name': file_name, 'sheet_name': selected_sheet, 'processor': processor.__name__})
    return result


def merge_summary(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    summary: Dict[str, Any] = {'total_rows': 0, 'valid_rows': 0, 'duplicate_rows': 0, 'errors_count': 0, 'skipped_rows': 0}
    for result in results:
        for key, value in result.get('summary', {}).items():
            if isinstance(value, int):
                summary[key] = summary.get(key, 0) + value
    return summary


def process_excel_import(
    uploaded_file,
    sheet_name: str = '',
    commit: bool = False,
    user=None,
    employee_import_mode: str = '',
    field_mapping: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    file_name = getattr(uploaded_file, 'name', 'uploaded.xlsx')
    requested = clean_text(sheet_name)
    normalized_employee_mode = normalize_header(employee_import_mode)
    if normalized_employee_mode not in ('', 'template', 'flexible'):
        raise ValueError('Employee import mode must be template or flexible.')
    importable_sheets = [sheet for sheet in workbook.sheetnames if resolve_processor(sheet) is not import_generic_sheet or sheet == 'Database']

    if normalized_employee_mode:
        if requested and requested not in workbook.sheetnames:
            raise ValueError(f'The requested worksheet "{requested}" was not found in the uploaded file.')
        if normalized_employee_mode == 'template':
            selected_sheet = requested or ('Employees' if 'Employees' in workbook.sheetnames else workbook.sheetnames[0])
        else:
            selected_sheet = requested or workbook.sheetnames[0]
        result = process_single_sheet(
            workbook,
            selected_sheet,
            commit,
            file_name,
            user=user,
            force_employee_roster=True,
            employee_import_mode=normalized_employee_mode,
            field_mapping=field_mapping,
        )
        result['available_sheets'] = workbook.sheetnames
        result['importable_sheets'] = workbook.sheetnames
        result['privacy_note'] = 'Raw Excel files are not stored in GitHub. Records are saved to PostgreSQL only when commit mode is selected.'
        return result

    if normalize_header(requested) in ('all', '__all__', 'كل الشيتات', 'all sheets'):
        results = [process_single_sheet(workbook, sheet, commit, file_name, user=user) for sheet in importable_sheets]
        summary = merge_summary(results)
        return {'mode': 'commit' if commit else 'preview', 'file_name': file_name, 'sheet_name': 'ALL', 'available_sheets': workbook.sheetnames, 'importable_sheets': importable_sheets, 'summary': summary, 'sheet_results': [{'sheet_name': item['sheet_name'], 'processor': item['processor'], 'summary': item['summary'], 'errors': item.get('errors', [])[:20], 'duplicates': item.get('duplicates', [])[:20], 'preview_rows': item.get('preview_rows', [])[:5]} for item in results], 'duplicates': [dup for item in results for dup in item.get('duplicates', [])][:50], 'errors': [err for item in results for err in item.get('errors', [])][:50], 'preview_rows': [row for item in results for row in item.get('preview_rows', [])][:12], 'privacy_note': 'Raw Excel files are not stored in GitHub. Records are saved to PostgreSQL only when commit mode is selected.'}

    if requested and requested not in workbook.sheetnames:
        raise ValueError(f'The requested worksheet "{requested}" was not found in the uploaded file.')
    selected_sheet = requested if requested in workbook.sheetnames else ('Database' if 'Database' in workbook.sheetnames else workbook.sheetnames[0])
    result = process_single_sheet(workbook, selected_sheet, commit, file_name, user=user)
    result['available_sheets'] = workbook.sheetnames
    result['importable_sheets'] = importable_sheets
    result['privacy_note'] = 'Raw Excel files are not stored in GitHub. Records are saved to PostgreSQL only when commit mode is selected.'
    return result
